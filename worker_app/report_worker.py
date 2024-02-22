import datetime
from django.http import JsonResponse, HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from django.db.models import Sum
from openpyxl.utils import get_column_letter
from rest_framework import status

from worker_app.models import Object, Shift, WorkType, WorkersBenefits, Travel, TravelBenefits, Worker


class ReportWorkerView(View):
    def get(self, request, *args, **kwargs):
        object_id = kwargs.get("object_id", None)

        # Получаем объект
        selected_object = Object.objects.filter(id=object_id).first()
        work_types = WorkType.objects.filter(category__object=selected_object)

        # Получаем все смены для данного объекта
        shifts = Shift.objects.filter(work_type__in=work_types)

        if not shifts.count():
            return HttpResponse({'Для данного объекта не обнаружено ни одной смены'}, status=status.HTTP_404_NOT_FOUND)

        # Получаем все командировки для данного объекта
        travels = Travel.objects.filter(object=selected_object)

        # Получаем список всех работников
        workers_from_travels = {travel.worker for travel in travels}
        workers_from_shifts = {shift.worker for shift in shifts}
        workers = workers_from_travels.union(workers_from_shifts)

        # Создаем новый Excel файл
        workbook = Workbook()

        for worker in workers:
            # Create a new worksheet for each worker
            worksheet = workbook.create_sheet(title=f'{worker.surname}_{worker.name}')

            # Названия колонок
            columns = ['Название работ', 'ед. изм.', 'кол-во', 'цена', 'сумма']

            # Добавляем строку с фамилией и именем рабочего над заголовками
            worker_name_row = worksheet.cell(row=1, column=1, value=f'{worker.surname}  {worker.name}')
            worker_name_row.font = Font(bold=True)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1,
                                  end_column=5)  # Объединяем ячейки на 5 колонок
            worksheet['A1'].alignment = Alignment(horizontal='center')  # Align the text in the merged cell to center

            # Записываем заголовки в файл для смен
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=2, column=col_num)  # Начинаем с новой строки (2)
                cell.value = column_title
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill
                cell.font = Font(bold=True)  # Make the text bold

                # Adding thin borders
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                cell.border = thin_border

            # # Устанавливаем ширину колонок
            worksheet.column_dimensions[get_column_letter(1)].width = 40
            worksheet.column_dimensions[get_column_letter(2)].width = 10
            worksheet.column_dimensions[get_column_letter(3)].width = 10
            worksheet.column_dimensions[get_column_letter(4)].width = 20
            worksheet.column_dimensions[get_column_letter(5)].width = 20

            # Записываем данные для выполненных работ
            row_num = 3
            total_amount = 0

            for work_type in WorkType.objects.all():
                shifts_filtered = shifts.filter(worker=worker, work_type=work_type)
                if shifts_filtered.exists():
                    total_scope = shifts_filtered.aggregate(Sum('value'))['value__sum']
                    price_for_worker = work_type.price_for_worker
                    total_price = total_scope * price_for_worker

                    worksheet.cell(row=row_num, column=1, value=work_type.name)
                    worksheet.cell(row=row_num, column=2, value=work_type.measurement_type.name).alignment = Alignment(
                        horizontal='center')
                    worksheet.cell(row=row_num, column=3, value=total_scope).alignment = Alignment(horizontal='center')
                    worksheet.cell(row=row_num, column=4, value=price_for_worker)
                    worksheet.cell(row=row_num, column=5, value=total_price)

                    total_amount += total_price
                    row_num += 1

            # Расчет выплат и остатков
            payments = WorkersBenefits.objects.filter(worker=worker, object=selected_object)
            total_payments = payments.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
            remaining_balance = total_amount - total_payments

            # Добавляем строки для выплат и остатков
            row_num += 1
            worksheet.cell(row=row_num, column=1, value='ИТОГО')
            worksheet.cell(row=row_num, column=5, value=total_amount)

            row_num += 1
            worksheet.cell(row=row_num, column=1, value='ВЫПЛАЧЕНО')
            worksheet.cell(row=row_num, column=5, value=total_payments)

            row_num += 1
            worksheet.cell(row=row_num, column=1, value='ОСТАТКИ К ВЫПЛАТЕ')
            worksheet.cell(row=row_num, column=5, value=remaining_balance)

            columns_travel = ['Рабочий', 'дней', 'к оплате', 'выплачено', 'остаток']

            # Заголовок КОМАНДИРОВОЧНЫЕ
            row_num += 4
            travels_row = worksheet.cell(row=row_num, column=1, value='Командировочные')
            travels_row.font = Font(bold=True, )
            worksheet.merge_cells(start_row=row_num,
                                  start_column=1,
                                  end_row=row_num,
                                  end_column=5)  # Объединяем ячейки на 5 колонок
            # Center-align the text in the merged cells
            travels_row.alignment = Alignment(horizontal='center', vertical='center')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            travels_row.fill = orange_fill

            # Записываем заголовки в файл для командировок
            row_num += 1
            for col_num, column_title in enumerate(columns_travel, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill
                cell.font = Font(bold=True)  # Make the text bold

                # Adding thin borders
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                cell.border = thin_border

            # Записываем данные для командировок
            row_num += 2
            total_remaining_balance = 0
            for travel in travels.filter(worker=worker):
                # Calculate days in travel
                finish = travel.date_finish or datetime.date.today()
                days_in_travel = (finish - travel.date_start).days + 1

                worksheet.cell(row=row_num, column=1,
                               value=str(travel.worker) + "   " + travel.date_start.strftime('%B'))  # Worker + Month
                worksheet.cell(row=row_num, column=2, value=days_in_travel)  # Days in travel
                worksheet.cell(row=row_num, column=3,
                               value=days_in_travel * travel.rate)  # Total for the month
                worksheet.cell(row=row_num, column=4,
                               value=TravelBenefits.objects.filter(travel=travel).aggregate(Sum('paid_for_travel'))[
                                         'paid_for_travel__sum'] or 0)  # Paid

                # Remaining balance
                remaining_balance = (days_in_travel * travel.rate) - (
                        TravelBenefits.objects.filter(travel=travel).aggregate(Sum('paid_for_travel'))[
                            'paid_for_travel__sum'] or 0)
                worksheet.cell(row=row_num, column=5, value=remaining_balance)  # Remaining balance
                total_remaining_balance += remaining_balance

                # Пустая строка после каждой записи
                row_num += 1

            # Добавляем строки для выплат и остатков
            row_num += 1
            worksheet.cell(row=row_num, column=1, value='ОСТАТКИ К ВЫПЛАТЕ')
            worksheet.cell(row=row_num, column=5, value=total_remaining_balance)

        # Удаляем первый лист тк он пустой (не понял где создаётся пока)
        first_sheet = workbook.sheetnames[0]
        workbook.remove(workbook[first_sheet])

        # Сохраняем файл
        report_path = f'report_worker.xlsx'
        workbook.save(report_path)

        # Отправляем файл пользователю
        with open(report_path, 'rb') as excel:
            response = HttpResponse(excel.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report_worker.xlsx'

        return response
