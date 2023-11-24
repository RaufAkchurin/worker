from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.db.models import Sum

from worker_app.models import Object, Shift, WorkType, WorkersBenefits


# TODO исключить 500 если нет данных для отчёта (несущ айди )

def generate_report(request, object_id):
    # Получаем объект
    selected_object = Object.objects.get(id=object_id)

    work_types = WorkType.objects.filter(category__object=selected_object)

    # Получаем все смены для данного объекта
    shifts = Shift.objects.filter(work_type__in=work_types)

    # Создаем новый Excel файл
    workbook = Workbook()
    worksheet = workbook.active

    # Названия колонок
    columns = ['Название типа работ', 'Тип измерения', 'Количество выполненного', 'Цена для рабочего', 'Сумма']

    # Записываем заголовки в файл
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal='center')

    # Записываем данные
    row_num = 2
    total_amount = 0

    for work_type in WorkType.objects.all():
        shifts_filtered = shifts.filter(work_type=work_type)
        if shifts_filtered.exists():
            total_scope = shifts_filtered.aggregate(Sum('value'))['value__sum']
            price_for_worker = work_type.price_for_worker
            total_price = total_scope * price_for_worker

            worksheet.cell(row=row_num, column=1, value=work_type.name)
            worksheet.cell(row=row_num, column=2, value=work_type.measurement_type.name)
            worksheet.cell(row=row_num, column=3, value=total_scope)
            worksheet.cell(row=row_num, column=4, value=price_for_worker)
            worksheet.cell(row=row_num, column=5, value=total_price)

            total_amount += total_price
            row_num += 1

    # Расчет выплат и остатков
    payments = WorkersBenefits.objects.filter(object=selected_object)
    total_payments = payments.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
    remaining_balance = total_amount - total_payments

    # Добавляем строки для выплат и остатков
    row_num += 1
    worksheet.cell(row=row_num, column=1, value='ВЫПЛАЧЕНО')
    worksheet.cell(row=row_num, column=5, value=total_payments)

    row_num += 1
    worksheet.cell(row=row_num, column=1, value='ОСТАТКИ К ВЫПЛАТЕ')
    worksheet.cell(row=row_num, column=5, value=remaining_balance)

    # Сохраняем файл
    report_path = f'report_workers/report_{selected_object.name}.xlsx'
    workbook.save(report_path)

    # Возвращаем путь к файлу
    return JsonResponse({'report_path': report_path})
