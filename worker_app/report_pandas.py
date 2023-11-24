from django.http import HttpResponse
from django.views import View
import pandas as pd
from openpyxl.styles import NamedStyle, Alignment, PatternFill, Font
from django.db.models import F
from openpyxl.utils import get_column_letter
from rest_framework import status

from .models import Shift, WorkType, Object


class GenerateReportView(View):
    def get(self, request, *args, **kwargs):
        # Извлекаем данные из базы
        object_id = kwargs.get("object_id", None)
        object = Object.objects.filter(id=object_id).first()
        work_types = WorkType.objects.filter(category__object=object)

        if not work_types.count():
            return HttpResponse({'WorkType for this object not found'}, status=status.HTTP_404_NOT_FOUND)

        work_types = work_types.annotate(
            сумма=F('total_scope') * F('price_for_customer'),
        ).values(
            'name',
            'category__name',  # Include the category name for grouping
            'measurement_type__name',
            'total_scope',
            'сумма',
        )

        # Преобразуем queryset в DataFrame
        df = pd.DataFrame(list(work_types))

        # Получаем данные из модели Shift и объединяем их с DataFrame
        shift_df = Shift.objects.values(
            work_type_name=F('work_type__name'),
            category_name=F('work_type__category__name'),  # Include the category name for grouping
            quantity_shift=F('value'),
            price_for_customer=F('work_type__price_for_customer'),
        )

        # Преобразуем queryset в DataFrame
        shift_df = pd.DataFrame(list(shift_df))

        # Объединяем DataFrames
        merged_df = pd.merge(df, shift_df, how='left', left_on='name', right_on='work_type_name')

        # Rename columns after the merge
        merged_df = merged_df.rename(columns={
            'name': 'Наименование работ',
            'category__name': 'Категория',  # Rename the category column
            'measurement_type__name': 'ед.изм.',
            'total_scope': 'кол-во',
            'price_for_customer': 'цена',
            'quantity_shift': 'кол-во выполненное',
        })

        # Вычисляем сумму для заказчика на оплату
        merged_df['сумма_заказчика'] = merged_df['кол-во выполненное'] * merged_df['цена']

        # Группируем данные по категории и типу работ, агрегируем значения
        grouped_df = merged_df.groupby(['Категория', 'Наименование работ']).agg({
            'ед.изм.': 'first',
            'кол-во': 'first',
            'цена': 'first',
            'сумма': 'sum',
            'кол-во выполненное': 'sum',
            'сумма_заказчика': 'sum',
        }).reset_index()

        # Calculate the overall total amount for the customer across all categories and work types
        overall_total_customer_amount = grouped_df['сумма_заказчика'].sum()

        # Создаем Excel-файл
        excel_file_path = 'report.xlsx'
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            # Создаем пустой лист 'WorkTypes'
            writer.book.create_sheet('WorkTypes')
#####################################################################################################################
            # Object  info

            # Записываем название объекта в первую строку и объединяем первые шесть ячеек
            object_name_row = pd.DataFrame({'Наименование объекта': [object.name] + [''] * 5})
            writer.sheets['WorkTypes'].append(list(object_name_row.iloc[0]), )

            # Устанавливаем высоту строки
            start_row = writer.sheets['WorkTypes'].max_row
            writer.sheets['WorkTypes'].row_dimensions[
                start_row].height = 25  # устанавливаем высоту в 30 пунктов

            # Объединяем первые шесть ячеек в первой строке
            start_col = 1
            end_col = 8
            end_row = start_row
            writer.sheets['WorkTypes'].merge_cells(start_row=start_row, start_column=start_col, end_row=end_row,
                                                   end_column=end_col)

            # Устанавливаем выравнивание для объединенных ячеек
            merged_cells = writer.sheets['WorkTypes'].cell(row=start_row, column=start_col)
            merged_cells.alignment = Alignment(horizontal='center', vertical='center')

            # Устанавливаем цвет для ячейки с названием объекта
            object_name_cell = writer.sheets['WorkTypes'].cell(row=start_row, column=start_col)
            object_name_cell.fill = PatternFill(start_color='808080', end_color='808080',
                                                fill_type='solid')  # Серый цвет

#####################################################################################################################

            # Продолжаем существующий цикл для записи данных категорий
            for category, category_df in grouped_df.groupby('Категория'):

                # Записываем название категории в каждую строку
                category_row = pd.DataFrame({'Наименование работ': [f'Категория: {category}'], 'Категория': ['']})
                writer.sheets['WorkTypes'].append(list(category_row.iloc[0]), )

                # Смержим ячейки и установим текст по центру
                merged_cell_range = f'A{writer.sheets["WorkTypes"].max_row}:H{writer.sheets["WorkTypes"].max_row}'
                writer.sheets['WorkTypes'].merge_cells(merged_cell_range)

                # Объединяем ячейки, делаем текст жирным, центрируем, устанавливаем желтый фон для названия категории
                for cell in list(writer.sheets['WorkTypes'].rows)[-1]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = NamedStyle(name='Bold').font.__class__(bold=True)
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Желтый цвет

                # Записываем данные для каждой категории
                category_df.to_excel(writer, sheet_name='WorkTypes', index=False,
                                     startrow=writer.sheets['WorkTypes'].max_row)
###################################################################################################################
                # итоговые данные после списка типов работ

                # Добавляем пустую строку
                category_row = pd.DataFrame({'Наименование работ': [f'итого: {category}'], })
                writer.sheets['WorkTypes'].append(list(category_row.iloc[0]), )

                # Рассчитываем и добавляем итог для каждой категории в колонку "summa_zakazchika"
                sum_customer_value = category_df['сумма_заказчика'].sum()
                sum_customer_row = pd.DataFrame({'сумма_заказчика': [sum_customer_value]})
                column_letter_1 = get_column_letter(writer.sheets['WorkTypes'].max_column)
                writer.sheets['WorkTypes'][
                    f'{column_letter_1}{writer.sheets["WorkTypes"].max_row + 1}'] = sum_customer_value

                # Рассчитываем и добавляем итог для каждой категории в колонку "summa"
                sum_value = category_df['сумма'].sum()
                sum_row = pd.DataFrame({'сумма': [sum_value]})
                column_letter_2 = get_column_letter(writer.sheets['WorkTypes'].max_column - 2)
                writer.sheets['WorkTypes'][
                    f'{column_letter_2}{writer.sheets["WorkTypes"].max_row}'] = sum_value

                # Определите адрес ячейки, которую вы хотите выделить жирным шрифтом
                cell_address_sum_customer = f'{column_letter_1}{writer.sheets["WorkTypes"].max_row}'
                cell_address_sum = f'{column_letter_2}{writer.sheets["WorkTypes"].max_row}'

                # Установите стиль для жирного шрифта
                bold_font = Font(bold=True)

                # Примените стиль к выбранным ячейкам
                writer.sheets['WorkTypes'][cell_address_sum_customer].font = bold_font
                writer.sheets['WorkTypes'][cell_address_sum].font = bold_font

            # Добавляем строку с расчётом итоговой суммы по всему объекту для заказчика

            column_letter_3 = get_column_letter(writer.sheets['WorkTypes'].max_column)
            writer.sheets['WorkTypes'][
                f'{column_letter_3}{writer.sheets["WorkTypes"].max_row + 2}'] = overall_total_customer_amount

            column_letter_4 = get_column_letter(writer.sheets['WorkTypes'].max_column - 6)
            writer.sheets['WorkTypes'][
                f'{column_letter_4}{writer.sheets["WorkTypes"].max_row}'] = "По объекту"

            # Определите бирюзовый цвет
            turquoise_fill = PatternFill(start_color='00FFEE', end_color='00FFEE', fill_type='solid')

            # Получите последнюю строку
            last_row = writer.sheets['WorkTypes'].max_row

            # Переберите ячейки в последней строке и установите цвет
            for col_num in range(1, writer.sheets['WorkTypes'].max_column + 1):
                cell = writer.sheets['WorkTypes'].cell(row=last_row, column=col_num)
                cell.fill = turquoise_fill


###################################################################################################################

            # Устанавливаем ширину для объединенных ячеек
            for col in range(start_col, end_col + 1):
                col_letter = get_column_letter(col)
                writer.sheets['WorkTypes'].column_dimensions[col_letter].width = 15

            # Скрытие первой колонки
            writer.sheets['WorkTypes'].column_dimensions['A'].hidden = True

            # Установка ширин для второй, третьей и четвертой колонок
            column_widths = {'A': 50, 'B': 57, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 20, 'H': 20}

            for col_letter, width in column_widths.items():
                writer.sheets['WorkTypes'].column_dimensions[col_letter].width = width

        # Отправляем файл пользователю
        with open(excel_file_path, 'rb') as excel:
            response = HttpResponse(excel.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report.xlsx'

        return response
